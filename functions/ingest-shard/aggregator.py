import io
import openpyxl
from datetime import datetime
from storage_adapter import DriveStorageAdapter

class InventoryAggregator:
    def __init__(self, drive_adapter: DriveStorageAdapter):
        self.drive = drive_adapter
        self.filename = "Kintsu_Inventory.xlsx"

    def ensure_inventory_file(self, folder_id: str) -> str:
        """
        Ensures the inventory file exists. Returns its ID.
        """
        file = self.drive.find_file_by_name(self.filename, folder_id)
        if file:
            return file['id']
        
        # Create new
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        # Header
        ws.append(["Date", "Item Name", "Category", "Merchant", "Total Amount", "Currency", "Confidence", "Source File"])
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        new_file = self.drive.create_file(
            name=self.filename,
            parent_id=folder_id,
            content=buffer.getvalue(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return new_file['id']

    def append_item(self, folder_id: str, item_data: dict, source_filename: str):
        file_id = self.ensure_inventory_file(folder_id)
        
        # Download
        content_bytes = self.drive.download_file(file_id)
        
        # Load Workbook
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content_bytes))
            ws = wb.active
        except Exception as e:
            print(f"Error loading Excel file: {e}. Creating new one.")
            # Fallback if file is corrupted
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inventory"
            ws.append(["Date", "Item Name", "Category", "Merchant", "Total Amount", "Currency", "Confidence", "Source File"])

        # Prepare row
        # item_data is the JSON from Gemini
        row = [
            item_data.get('date'),
            item_data.get('item_name'),
            item_data.get('category'),
            item_data.get('merchant'),
            item_data.get('total_amount'),
            item_data.get('currency'),
            item_data.get('confidence'),
            source_filename
        ]
        
        ws.append(row)
        
        # Save
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Update
        self.drive.update_file(
            file_id=file_id,
            content=buffer.getvalue(),
            mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        print(f"Appended item to {self.filename}")
